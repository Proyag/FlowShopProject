# Use this for GUI version
import numpy as np
import openpyxl as ox
import sys

from itertools import permutations
from tkinter import ttk, font
from tkinter import Tk, N, E, W, S, IntVar, StringVar

from utils import *
import genetic


taillard_file = ox.load_workbook('Makespans.xlsx')  # Load workbook
taillard = taillard_file.get_sheet_by_name('Sheet1')  # Load sheet


def calculate_optimal_makespan(*args):
    # To make command line print complete nd arrays
    np.set_printoptions(threshold=sys.maxsize)

    problem = problem_num.get()

    jobs = taillard.cell(row=problem+3, column=1).value
    machines = taillard.cell(row=problem+3, column=2).value
    timeseed = taillard.cell(row=problem+3, column=3).value

    # Write these to GUI
    jobs_val.set(jobs)
    macs_val.set(machines)
    time_val.set(timeseed)

    # Generate matrix of times
    a = random_matrix(machines, jobs, timeseed)
    print("Problem No.:", problem, "\nMatrix:\n", a)
    # Transpose matrix to calulate makespan
    at = a.transpose()

    # Start with all possible permutations of first 4 jobs
    init_jobs = 4
    init_job_list = list(range(init_jobs))
    # Generate all 24 permutations
    sequence_list = np.array(list(permutations(init_job_list)))

    # This loop will eventually run till init_jobs <= jobs
    while init_jobs <= jobs:
        makespan_list = np.array([])
        for seq in sequence_list:
            seq = list(seq)
            makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], seq))

        # Sort both arrays in ascending order of makespan
        # and reduce to 20 best sequences
        sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

        # Roulette wheel .. 3 x 20 output
        sequence_list, makespan_list = genetic.roulette_wheel(sequence_list, makespan_list)

        # Again, sort and reduce to 20
        sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

        # Now ordered crossover each of the 20 sequences
        # with each other and add to lists
        for i in range(20):
            for j in range(20):
                if i != j:
                    child = genetic.ordered_crossover(sequence_list[i], sequence_list[j])
                    child = np.array(child)
                    sequence_list = np.vstack((sequence_list, child))
                    makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(child)))

        # Sort both arrays in ascending order of makespan
        # and reduce to 20 best sequences
        sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

        # Applying mutation. Inverse Mutation.
        for i in sequence_list:
            mutated = genetic.inverse_mutation(i)
            mutated = np.array(mutated)
            sequence_list = np.vstack((sequence_list, mutated))
            makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(mutated)))

        # Sort both arrays in ascending order of makespan
        # and reduce to 20 best sequences
        sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

        # Applying another mutation. Pairwise Swap Mutation.
        for i in sequence_list:
            mutated = genetic.pairwise_swap_mutation(i)
            mutated = np.array(mutated)
            sequence_list = np.vstack((sequence_list, mutated))
            makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(mutated)))

        # Sort both arrays in ascending order of makespan
        # and reduce to 20 best sequences
        sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

        # Ordered Crossover again
        for i in range(20):
            for j in range(20):
                if i != j:
                    child = genetic.ordered_crossover(sequence_list[i], sequence_list[j])
                    child = np.array(child)
                    sequence_list = np.vstack((sequence_list, child))
                    makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(child)))

        # Sort both arrays in ascending order of makespan
        # and reduce to 20 best sequences
        sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

        # Set first element of sorted list as best makespan.
        best_sequence = sequence_list[0]
        best_makespan = makespan_list[0]

        # Bring in next job into every position in sequence_list
        init_jobs += 1
        init_job_list = list(range(init_jobs))
        new_sequence_list = np.array([], dtype=int).reshape(0, init_jobs)
        for s in sequence_list:
            for pos in range(s.size + 1):
                new_sequence_list = np.vstack((new_sequence_list, np.insert(s, pos, init_jobs - 1)))
        sequence_list = new_sequence_list

    print("\nBest Sequence:\n", best_sequence, "\n\n")
    makespan_val.set(best_makespan)


def increment(*args):
    problem_num.set(problem_num.get() + 1)
    calculate_optimal_makespan()


def decrement(*args):
    problem_num.set(problem_num.get() - 1)
    calculate_optimal_makespan()


if __name__ == '__main__':
    root = Tk()
    # Format window size
    # root.geometry('{}x{}'.format(1920, 1080))
    # Give the window a name
    root.title("Scheduling Optimization")

    # Creating frame inside window for ttk stuff
    mainframe = ttk.Frame(root, padding="3 3 12 12")
    mainframe.grid(column=0, row=0, sticky=(N, S, E, W))
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    mainframe.columnconfigure(1, weight=1)
    mainframe.columnconfigure(2, weight=3)
    mainframe.columnconfigure(3, weight=1)
    mainframe.rowconfigure(1, weight=1)
    mainframe.rowconfigure(2, weight=1)
    mainframe.rowconfigure(3, weight=1)
    mainframe.rowconfigure(4, weight=1)
    mainframe.rowconfigure(5, weight=1)
    mainframe.rowconfigure(6, weight=1)
    mainframe.rowconfigure(7, weight=1)

    MyFont = font.Font(family='Helvetica', size=20)

    problem_num = IntVar()
    makespan_val = IntVar()
    jobs_val = IntVar()
    macs_val = IntVar()
    time_val = IntVar()

    problem_entry = ttk.Entry(mainframe, width=20, font=MyFont, textvariable=problem_num)
    problem_entry.grid(column=2, row=1, sticky=(W, E))
    ttk.Label(mainframe, textvariable=makespan_val, font=MyFont).grid(column=2, row=6, sticky=(W, E))
    ttk.Button(mainframe, text="Calculate", width=30, command=calculate_optimal_makespan).grid(column=3, row=1, sticky=(E))
    ttk.Label(mainframe, text="Problem No.", font=MyFont).grid(column=1, row=1, sticky=(W, E))
    ttk.Label(mainframe, text="Optimal makespan", font=MyFont).grid(column=1, row=6, sticky=(W, E))

    ttk.Label(mainframe, text="Jobs", font=MyFont).grid(column=1, row=3, sticky=(W, E))
    ttk.Label(mainframe, textvariable=jobs_val, font=MyFont).grid(column=2, row=3, sticky=(W, E))
    ttk.Label(mainframe, text="Machines", font=MyFont).grid(column=1, row=4, sticky=(W, E))
    ttk.Label(mainframe, textvariable=macs_val, font=MyFont).grid(column=2, row=4, sticky=(W, E))
    ttk.Label(mainframe, text="Timeseed", font=MyFont).grid(column=1, row=5, sticky=(W, E))
    ttk.Label(mainframe, textvariable=time_val, font=MyFont).grid(column=2, row=5, sticky=(W, E))

    calculating = StringVar()
    calculating.set("")
    ttk.Label(mainframe, textvariable=calculating).grid(column=3, row=5, sticky=(W, E))

    ttk.Button(mainframe, text="Next", width=30, command=increment).grid(column=3, row=7, sticky=E)
    ttk.Button(mainframe, text="Previous", width=30, command=decrement).grid(column=1, row=7, sticky=W)

    # Padding around every widget in the frame
    for child in mainframe.winfo_children():
        child.grid_configure(padx=40, pady=40)
    # Focus on the entry field at first
    problem_entry.focus()
    # Executes function when you press Enter
    root.bind('<Return>', calculate_optimal_makespan)

    # Changing theme
    s = ttk.Style()
    s.theme_use('clam')
    s.configure('TButton', font=MyFont)

    # Start the infinite loop
    root.mainloop()
