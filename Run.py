# Use this program to write to Excel
import numpy as np
import openpyxl as ox
import sys

from itertools import permutations

import genetic


# Read jobs, machines, timeseed columns from Taillard.xlsx and store them in three lists
# Initialize lists
timeseed_list = []
jobs_list = []
machines_list = []
makespans = []
taillard_file = ox.load_workbook('../Makespans.xlsx')  # Load workbook
taillard = taillard_file.get_sheet_by_name('Sheet1')  # Load sheet

# Values are in rows 3 to 122 of the Taillard.xlsx Excel sheet
for i in range(3, 123):
    j = taillard.cell(row=i, column=1).value  # Read Jobs value
    m = taillard.cell(row=i, column=2).value  # Read Machines value
    t = taillard.cell(row=i, column=3).value  # Read Timeseed value
    v = taillard.cell(row=i, column=6).value  # Read our best makespan
    jobs_list.append(j)
    machines_list.append(m)
    timeseed_list.append(t)
    makespans.append(v)


def main():
    # To make command line print complete nd arrays.
    np.set_printoptions(threshold=sys.maxsize)
    # Run for entire 120 problem range
    for problem in range(0, 120):
        jobs = jobs_list[problem]
        machines = machines_list[problem]
        timeseed = timeseed_list[problem]
        # Generate processing time matrix
        a = random_matrix(machines, jobs, timeseed)

        # Transpose matrix to calulate makespan
        at = a.transpose()

        # Start with all possible permutations of first 4 jobs
        init_jobs = 4
        init_job_list = list(range(init_jobs))
        # Generate all 24 permutations
        sequence_list = np.array(list(permutations(init_job_list)))

        # This loop will run till init_jobs == jobs
        while init_jobs <= jobs:
            makespan_list = np.array([])
            for seq in sequence_list:
                seq = list(seq)
                makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], seq))

            # Sort both arrays in ascending order of makespan
            # and reduce to 20 best sequences
            sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

            # Roulette wheel .. 3 x 20 output
            sequence_list, makespan_list = genetic.roulette_wheel(sequence_list, makespan_list)

            # Again, sort and reduce to 20
            sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)
            # print("Best 20\n", sequence_list, makespan_list)

            # Now ordered crossover each of the 20 sequences with each other and add to lists
            for i in range(20):
                for j in range(20):
                    if i != j:
                        child = genetic.ordered_crossover(sequence_list[i], sequence_list[j])
                        child = np.array(child)
                        sequence_list = np.vstack((sequence_list, child))
                        makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(child)))

            # Sort both arrays in ascending order of makespan
            # and reduce to 20 best sequences
            sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

            # Applying mutation. Inverse mutation.
            for i in sequence_list:
                mutated = genetic.inverse_mutation(i)
                mutated = np.array(mutated)
                sequence_list = np.vstack((sequence_list, mutated))
                makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(mutated)))

            # Sort both arrays in ascending order of makespan
            # and reduce to 20 best sequences
            sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

            # Applying another mutation. Pairwise Swap Mutation.
            for i in sequence_list:
                mutated = genetic.pairwise_swap_mutation(i)
                mutated = np.array(mutated)
                sequence_list = np.vstack((sequence_list, mutated))
                makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(mutated)))

            # Sort both arrays in ascending order of makespan
            # and reduce to 20 best sequences
            sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

            # Ordered Crossover again
            for i in range(20):
                for j in range(20):
                    if i != j:
                        child = genetic.ordered_crossover(sequence_list[i], sequence_list[j])
                        child = np.array(child)
                        sequence_list = np.vstack((sequence_list, child))
                        makespan_list = np.append(makespan_list, calculate_makespan(at[init_job_list], list(child)))

            # Sort both arrays in ascending order of makespan
            # and reduce to 20 best sequences
            sequence_list, makespan_list = sort_and_reduce(sequence_list, makespan_list)

            # Set first element of sorted list as best makespan.
            best_sequence = sequence_list[0]
            best_makespan = makespan_list[0]

            # Bring in next job into every position in sequence_list
            init_jobs += 1
            init_job_list = list(range(init_jobs))
            new_sequence_list = np.array([], dtype=int).reshape(0, init_jobs)
            for s in sequence_list:
                for pos in range(s.size + 1):
                    new_sequence_list = np.vstack((new_sequence_list, np.insert(s, pos, init_jobs - 1)))
            sequence_list = new_sequence_list

        best_sequence = list(map(int, best_sequence))
        best_sequence = list(map(str, best_sequence))
        best_sequence = ', '.join(best_sequence)
        if best_makespan < makespans[problem]:
            taillard.cell(row=problem + 3, column=6).value = best_makespan
            taillard.cell(row=problem + 3, column=7).value = best_sequence
            taillard_file.save('Makespans.xlsx')
        print("Jobs: ", jobs, "\nMachines: ", machines, "\nTimeseed: ", timeseed, "\nOptimal sequence: ", best_sequence, "\nOptimal makespan: ", best_makespan, "\n")


if __name__ == '__main__':
        main()
